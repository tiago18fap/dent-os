"""Adiciona aba '13. crt.sh + LinkedIn' com achados da rodada 2 (Chrome browsing)."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/sessions/dreamy-brave-keller/mnt/GERADO PELO COWORK/EasyDental_v2.xlsx"
wb = load_workbook(XLSX)

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
HIGH_FILL = PatternFill("solid", start_color="C6EFCE")
MED_FILL = PatternFill("solid", start_color="FFEB9C")
LOW_FILL = PatternFill("solid", start_color="FFC7CE")
GOLD_FILL = PatternFill("solid", start_color="FFD966")
WHITE_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="FFFFFF")
BOLD = Font(name="Arial", bold=True, size=10)
NORMAL = Font(name="Arial", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

if "13. Rodada 2 - crt.sh+LinkedIn" in wb.sheetnames:
    del wb["13. Rodada 2 - crt.sh+LinkedIn"]
ws = wb.create_sheet("13. Rodada 2 - crt.sh+LinkedIn")

ws.merge_cells("A1:H1")
c = ws["A1"]
c.value = "RODADA 2 — Coleta via crt.sh + LinkedIn + Instagram (Chrome)"
c.fill = HEADER_FILL
c.font = TITLE_FONT
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

# Resultado crt.sh
ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = ("RESULTADO crt.sh: o EasyDental usa CERTIFICADO WILDCARD (*.easydentalcloud.com.br). "
           "Isso esconde os tenants — só 26 subdomínios apareceram, e a maioria é infra (mx, ftp, pop, "
           "tst-*). Subdomínios de clientes (ex: sescdf.easydentalcloud.com.br) NÃO aparecem no CT log. "
           "Conclusão: crt.sh é DEAD-END para descobrir clientes — precisamos do LinkedIn + Instagram + vagas.")
c.fill = LOW_FILL
c.font = Font(name="Arial", bold=True, size=10, color="000000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws.row_dimensions[2].height = 60

headers = ["#", "Categoria", "Nome / Handle", "Local", "Cargo / Contexto",
           "Evidência", "Fonte", "Próxima ação"]
for j, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=j, value=h)
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER

novos = [
    # === CLÍNICAS NOVAS IDENTIFICADAS ===
    (1, "CLÍNICA — confirmar exato",
     "Sorrir Odontologia",
     "Florianópolis - SC (provável) — múltiplos candidatos: clinicaodontologicasorrir, Studio Sorrir Recife, Sorrir Belém",
     "Clínica",
     "Post no Instagram: 'Aqui na Sorrir temos utilizado o software da @easydentalcloud'",
     "Instagram @clinicaodontologicasorrir + busca Google",
     "Confirmar qual unidade Sorrir é (Florianópolis tem 7.386 seguidores no IG)"),

    (2, "CONSULTÓRIO — confirmar uso atual",
     "Consultório Dr. Renato Schiavoni",
     "Ribeirão Preto - SP",
     "Dentista (5★ Doctoralia, especialista em endodontia)",
     "Foi marcado em post do @easydentalcloud no Instagram — 'drrenatoschiavoni's profile picture'",
     "doctoralia.com.br/renato-schiavoni/dentista/ribeirao-preto",
     "Ligar/IG para confirmar uso ativo"),

    (3, "REDE — investigar",
     "Patroni (Clínica ou rede odontológica)",
     "Confirmar UF",
     "Débora Silveira é Gerente",
     "LinkedIn de Débora Silveira (perfil aparece em busca 'easydental site:linkedin.com')",
     "Google snippet LinkedIn",
     "Buscar 'Patroni odontologia' + LinkedIn da Débora para localizar"),

    # === FUNCIONÁRIAS DE CLÍNICAS — usam EasyDental ===
    (4, "Funcionária — descobrir clínica",
     "Carmen Lopes",
     "Brasil (CRO/LinkedIn confirmar)",
     "Gerente administrativa — 'trabalhei por 8 anos em clínica odontológica, domínio EasyDental'",
     "Perfil LinkedIn aparece em busca 'easydental site:linkedin.com/in'",
     "linkedin.com (busca pública)",
     "Abrir perfil LinkedIn → identificar clínica empregadora"),

    (5, "Funcionária — descobrir clínica",
     "Bianca Pereira da Silva",
     "Brasil",
     "Recepcionista / Administrativo — cita EasyDental no perfil",
     "Perfil LinkedIn público",
     "linkedin.com",
     "Abrir perfil → mapear clínica atual"),

    (6, "Funcionária — descobrir clínica",
     "Barbara Monteiro",
     "Brasil",
     "Perfil cita EasyDental",
     "Resultado Google LinkedIn",
     "linkedin.com",
     "Abrir perfil → mapear clínica"),

    (7, "Funcionária — descobrir clínica",
     "Eliza Camile Patriniani",
     "Brasil",
     "Assistente de Marketing — cita EasyDental",
     "Resultado Google LinkedIn",
     "linkedin.com",
     "Abrir perfil → mapear clínica/empresa"),

    (8, "Funcionária — descobrir clínica",
     "Renata De Valentim",
     "Brasil",
     "Cirurgiã-Dentista (na Prefeitura) — cita EasyDental",
     "Resultado Google LinkedIn",
     "linkedin.com",
     "Confirmar — pode ser uso público municipal"),

    (9, "Funcionária — descobrir clínica",
     "Kristine Lopes Do Amaral Neves",
     "Brasil",
     "Dentista Clínico Geral — cita EasyDental",
     "Resultado Google LinkedIn",
     "linkedin.com",
     "Abrir perfil → mapear clínica"),

    (10, "Funcionária — descobrir clínica",
     "Victoria Farias",
     "Brasil",
     "Cirurgiã-Dentista — cita EasyDental",
     "Resultado Google LinkedIn",
     "linkedin.com",
     "Abrir perfil → mapear clínica"),

    (11, "Funcionária da EasyDental (não cliente)",
     "Maria Claudia Harada Ferreira",
     "São Carlos - SP",
     "Trabalha NA EasyDental (UNICAMP, 161 conexões)",
     "Perfil LinkedIn",
     "linkedin.com",
     "NÃO é cliente — funcionária do concorrente. Pular."),

    # === KOL adicional confirmado neste round ===
    (12, "KOL — confirmar uso",
     "Dr. Renato Schiavoni (acima) também aparece marcado",
     "Ribeirão Preto - SP",
     "Dentista que aparece em posts EasyDental Cloud",
     "Mesma evidência — post EasyDental Instagram",
     "instagram.com/easydentalcloud",
     "Já listado em #2 — duplicata"),

    # === PARCEIRO COMERCIAL (não é clínica) ===
    (13, "PARCEIRO comercial (NÃO cliente)",
     "Saúde Service",
     "Brasil (nacional)",
     "Provê máquina de cartão para dentistas — integrada com EasyDental Cloud",
     "Página oficial EasyDental + materiais.saudeservice.com.br/parcerias-easy-dental",
     "easydental.com.br/materiais/beneficios-da-parceria-saude-service-easydental/",
     "Não é clínica — é fornecedor parceiro. Mas a base de clientes Saúde Service tem alta sobreposição com EasyDental"),
]

for i, row in enumerate(novos, start=5):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = BORDER
    if (i - 5) % 2 == 1:
        for j in range(1, 9):
            ws.cell(row=i, column=j).fill = ALT_FILL

# Subdomínios crt.sh — bloco de infra
r0 = 5 + len(novos) + 2
ws.cell(row=r0, column=1, value="SUBDOMÍNIOS easydentalcloud.com.br ENCONTRADOS NO crt.sh (infra apenas)").font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws.cell(row=r0, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=8)

subdoms = [
    ("app.easydentalcloud.com.br", "Login da aplicação principal"),
    ("blog.easydentalcloud.com.br", "Blog"),
    ("dentalpartner.easydentalcloud.com.br", "Portal de parceiros"),
    ("ftp/mx1/mx2/pop/webmail.easydentalcloud.com.br", "E-mail e FTP"),
    ("lp.easydentalcloud.com.br", "Landing pages"),
    ("mobile.easydentalcloud.com.br", "Mobile (app fallback)"),
    ("notify.easydentalcloud.com.br + notify-tst", "Notificações"),
    ("portal.easydentalcloud.com.br + www.portal", "Portal informativo"),
    ("printer.easydentalcloud.com.br", "Servidor de impressão"),
    ("recursos.portal.easydentalcloud.com.br", "Recursos do portal"),
    ("semanadocliente.easydentalcloud.com.br", "Site de evento (Semana do Cliente)"),
    ("tst-clidec / tst-rct / tst-sesc / tst-sescdn / tst-sescsc", "Ambientes de TESTE — sesc, sescdn=SESC DF?, sescsc=SESC SC?"),
    ("videos.easydentalcloud.com.br", "Hospedagem de vídeos"),
    ("*.easydentalcloud.com.br", "Wildcard cert — cobre todos tenants"),
]
for i, (sd, desc) in enumerate(subdoms, start=r0+1):
    ws.cell(row=i, column=1, value=sd).font = NORMAL
    ws.cell(row=i, column=1).alignment = WRAP
    ws.cell(row=i, column=1).border = BORDER
    c = ws.cell(row=i, column=2, value=desc)
    c.font = NORMAL
    c.alignment = WRAP
    c.border = BORDER
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

# Pista importante: tst-sesc
r1 = r0 + len(subdoms) + 2
ws.merge_cells(start_row=r1, start_column=1, end_row=r1, end_column=8)
c = ws.cell(row=r1, column=1, value=(
    "💡 PISTA: os subdomínios 'tst-sesc', 'tst-sescdn', 'tst-sescsc' sugerem ambientes de teste para "
    "SESC SP, SESC DF/DN, SESC SC/Santa Catarina. Confirma a forte presença SESC no portfólio EasyDental "
    "(SESC SP + DF + RS + provável SC)."
))
c.font = Font(name="Arial", italic=True, size=10)
c.fill = GOLD_FILL
c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws.row_dimensions[r1].height = 50

# Próximos passos
r2 = r1 + 2
ws.cell(row=r2, column=1, value="PRÓXIMOS PASSOS DESTA RODADA").font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws.cell(row=r2, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=8)

passos = [
    "1. Abrir cada perfil LinkedIn listado acima (Carmen Lopes, Bianca, Barbara, Renata, Kristine, Victoria, Débora Silveira) e mapear a CLÍNICA empregadora atual. Cada perfil = 1 clínica nova confirmada.",
    "2. Buscar 'Patroni odontologia' para confirmar se é rede ou clínica única.",
    "3. Buscar Sorrir Odontologia + 'easydental' em cada candidata (Florianópolis, Belém, Recife) para descobrir qual usa.",
    "4. Solicitar via Instagram à @easydentalcloud a lista de cases recentes — eles postam mensalmente clínicas que usam.",
    "5. Abrir a página oficial easydental.com.br/depoimentos/ via Chrome (precisa JS) e extrair todos os nomes de clínicas/médicos depoentes.",
]
for i, p in enumerate(passos, start=r2+1):
    c = ws.cell(row=i, column=1, value=p)
    c.font = NORMAL
    c.alignment = WRAP
    c.border = BORDER
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    ws.row_dimensions[i].height = 32

widths = [4, 28, 30, 30, 32, 38, 30, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# Move tab to right position
order = wb.sheetnames
if "10. Fontes" in order:
    # mover 13 para antes da 10
    idx_13 = order.index("13. Rodada 2 - crt.sh+LinkedIn")
    idx_10 = order.index("10. Fontes")
    wb.move_sheet("13. Rodada 2 - crt.sh+LinkedIn", offset=idx_10 - idx_13 - 1)

wb.save(XLSX)
print(f"Saved: {XLSX}")
print(f"Aba '13. Rodada 2' adicionada com {len(novos)} novos achados + {len(subdoms)} subdomínios")
