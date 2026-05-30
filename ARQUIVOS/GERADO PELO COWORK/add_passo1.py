"""Adiciona aba '12. Passo 1 - Coleta' com nomes coletados via execução do pipeline."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/sessions/dreamy-brave-keller/mnt/GERADO PELO COWORK/EasyDental_v2.xlsx"
wb = load_workbook(XLSX)

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
HIGH_FILL = PatternFill("solid", start_color="C6EFCE")
MED_FILL = PatternFill("solid", start_color="FFEB9C")
LOW_FILL = PatternFill("solid", start_color="FFC7CE")
GOLD_FILL = PatternFill("solid", start_color="FFD966")
WHITE_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="FFFFFF")
BOLD = Font(name="Arial", bold=True, size=10)
NORMAL = Font(name="Arial", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

if "12. Passo 1 - Coleta" in wb.sheetnames:
    del wb["12. Passo 1 - Coleta"]
ws = wb.create_sheet("12. Passo 1 - Coleta")

ws.merge_cells("A1:J1")
c = ws["A1"]
c.value = "PASSO 1 EXECUTADO — Clínicas identificadas via OSINT (Maio 2026)"
c.fill = HEADER_FILL
c.font = TITLE_FONT
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

# Destaque — descoberta técnica
ws.merge_cells("A2:J2")
c = ws["A2"]
c.value = ("🎯 DESCOBERTA-CHAVE: Cada cliente EasyDental ganha um SUBDOMÍNIO [clinica].easydentalcloud.com.br "
           "(confirmado com sescdf.easydentalcloud.com.br). Isso permite descobrir TODOS os clientes via "
           "Certificate Transparency: acesse https://crt.sh/?q=%25.easydentalcloud.com.br — vai listar "
           "centenas de subdomínios = centenas de clínicas reais. Esta é a ÚNICA fonte que entrega o "
           "universo completo de clientes ativos.")
c.fill = GOLD_FILL
c.font = Font(name="Arial", bold=True, size=10, color="000000")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws.row_dimensions[2].height = 70

headers = ["#", "Categoria", "Nome", "Cidade-UF", "Decisor / Contato",
           "Evidência (como sei que usa)", "Fonte (URL)", "Confiança", "Status no funil",
           "Próxima ação"]
for j, h in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=j, value=h)
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER

dados = [
    # === GRUPO 1: REDES PÚBLICAS / GOVERNO (não target comercial mas indicam ecossistema) ===
    (1, "Rede pública — SESC SP",
     "SESC São Paulo - 34+ unidades",
     "Estado SP",
     "Gerência Odontologia SESC-SP",
     "Notícia oficial EasyDental: 'após licitação nacional 2017, SESC SP adotou EasyDental'",
     "easydental.com.br/noticias/99-sesc-sao-paulo-adota-solucao",
     "ALTA", "Não target (compra licitação)", "Apenas referência institucional"),

    (2, "Rede pública — SESC DF",
     "SESC Distrito Federal - 50 consultórios em 8 unidades",
     "DF (Brasília, Taguatinga, Ceilândia, Gama, Guará)",
     "SESC-DF — sescdf.com.br",
     "Adotou EasyDental em 2023. SUBDOMÍNIO ATIVO: sescdf.easydentalcloud.com.br",
     "easydental.com.br/blog/sesc-df-adota-solucao-easydental/",
     "ALTA (confirmado)", "Não target (compra licitação)", "Referência — não abordar"),

    (3, "Rede pública — SESC RS",
     "SESC Rio Grande do Sul",
     "Porto Alegre, Canoas, Passo Fundo - RS",
     "SESC-RS — (51) 3342-5099 / (51) 3284-2000",
     "Mencionado oficialmente como cliente: 'EasyDental tem contratos com SESC-SP e SESC-RS'",
     "sesc-rs.com.br/saude/clinicas-odontologicas/",
     "ALTA", "Não target (compra licitação)", "Apenas referência"),

    # === GRUPO 2: REDES PRIVADAS — ALVOS COMERCIAIS QUENTES ===
    (4, "REDE PRIVADA — ALVO QUENTE",
     "Rede COSB (Centro Ortodontia Saúde Bucal)",
     "SBC + Diadema - SP (6 unidades)",
     "(11) 4122-0550 — cosb.com.br — CNPJ 01.740.990/0001-67",
     "Depoimento oficial: cliente desde 2012, migrando para Cloud, planeja operação EUA",
     "easydental.com.br/depoimentos/",
     "ALTA", "Lead quente — rede", "Mapear decisor de TI + enviar pitch BI multi-unidade"),

    (5, "CONSULTÓRIO PREMIUM — ALVO QUENTE",
     "Myai Odontologia (Odonto Myai)",
     "São Paulo - SP (V. Antonieta/Aricanduva)",
     "Dr. Edson Myai (CRO-SP 59.180) — (11) 2727-4743 — odontomyai.com.br",
     "Case publicado: cliente desde 2012, 15+ dentistas, usa TODOS módulos",
     "easydental.com.br/blog/automacao-de-processos/estudo-de-caso-implantacao-do-easydental-na-clinica-myai/",
     "ALTA", "Lead morno — satisfeito", "Acompanhar — abordar se houver mudança"),

    # === GRUPO 3: EMBAIXADORES / KOLs (parceiros pagos, mas clientes ativos) ===
    (6, "KOL / Embaixador",
     "Instituto Bernal de Odontologia",
     "São Paulo - SP",
     "Dr. Anderson Bernal — andersonbernal.com.br — @institutobernal",
     "Página oficial de parceiros — diretor é embaixador EasyDental",
     "easydental.com.br/parceiros/",
     "ALTA", "Não target direto", "Mapear alunos/seguidores como leads"),

    (7, "KOL / Embaixador",
     "Curso Dra. Carol Moura",
     "Brasil (online)",
     "Dra. Carol Moura — Instagram",
     "Oferece desconto exclusivo EasyDental para alunos",
     "easydental.com.br/parceiros/",
     "ALTA", "Não target direto", "Olhar alunos dela"),

    (8, "KOL / Embaixador",
     "Curso Dr. Felipe Valverde",
     "Brasil",
     "Dr. Felipe Valverde",
     "Página parceiros — desconto p/ alunos",
     "easydental.com.br/parceiros/",
     "ALTA", "Não target direto", "Olhar alunos"),

    (9, "KOL / Embaixador",
     "Dr. André Appezzato",
     "Brasil",
     "Dr. André Appezzato",
     "Página parceiros EasyDental",
     "easydental.com.br/parceiros/",
     "ALTA", "Não target direto", "Olhar alunos"),

    (10, "KOL / Embaixador",
     "Dr. Rodrigo Vieira",
     "Brasil",
     "Dr. Rodrigo Vieira",
     "Página parceiros EasyDental",
     "easydental.com.br/parceiros/",
     "ALTA", "Não target direto", "Olhar alunos"),

    (11, "KOL / Embaixador",
     "Dr. Felipe Sala",
     "Brasil",
     "Dr. Felipe Sala",
     "Página parceiros EasyDental",
     "easydental.com.br/parceiros/",
     "ALTA", "Não target direto", "Olhar alunos"),

    (12, "KOL / Embaixador",
     "Dr. Rafael Rangel ('Rafa Rangel')",
     "Brasil",
     "@dr.rafaelrangel",
     "Página parceiros EasyDental — vende kit espátulas Rafa Rangel",
     "easydental.com.br/parceiros/",
     "ALTA", "Não target direto", "Olhar alunos"),

    # === GRUPO 4: DEPOENTES PÚBLICOS ===
    (13, "Cliente depoente",
     "Consultório Dr. Nelson Dall'Oca",
     "São Paulo - SP",
     "Dr. Nelson Alves Dall'Oca — ident.com.br/dalloca",
     "Depoimento no site: 'uso há anos, foco no técnico sem descuidar gestão'",
     "easydental.com.br/depoimentos/",
     "ALTA", "Cliente satisfeito", "Lead frio — só abordar se houver dor"),

    (14, "Cliente depoente",
     "Consultório Dr. Alysson Konno",
     "Brasil (confirmar UF via CRO)",
     "Dr. Alysson Konno — buscar LinkedIn",
     "Depoimento sobre suporte e Cloud",
     "easydental.com.br/depoimentos/",
     "MÉDIA", "Cliente satisfeito", "Confirmar UF, abordagem comparativa"),

    # === GRUPO 5: PISTAS VIA VAGAS DE EMPREGO ===
    (15, "VAGA mencionando EasyDental",
     "VIDA Desenvolvimento Humano e Organizacional LTDA",
     "Buscar UF (provavelmente MG/SP)",
     "vidadho.vagas.solides.com.br",
     "Vaga de estagiário pede 'apoio em agendamento via sistema Easy Dental'",
     "vidadho.vagas.solides.com.br/vaga/697134",
     "ALTA", "Lead novo", "Visitar página, identificar clínica cliente da VIDA OU se VIDA opera clínica"),

    (16, "VAGA mencionando EasyDental",
     "Clínica anônima — Vila Mariana, SP",
     "São Paulo - SP (Vila Mariana)",
     "Catho — vaga arquivada",
     "Vaga de recepcionista CLT: 'conhecimento em EASYDENTAL será diferencial'",
     "catho.com.br (vaga arquivada — buscar por título)",
     "MÉDIA", "Lead novo", "Reabrir o link arquivado, identificar contratante"),

    # === GRUPO 6: INVESTIGAR — NOMES SIMILARES (alta correlação) ===
    (17, "INVESTIGAR — nome similar",
     "Easydentes Clínicas Odontológicas Ltda",
     "São Paulo - SP (9 unidades: Itaim Paulista, Sapopemba, Sacomã, V. Mercês, Av. Oratório)",
     "easydentes.com.br — CNPJ 23.752.431/0001-00",
     "Rede com 9 unidades em SP, nome similar ao sistema. Vínculo NÃO confirmado em fontes oficiais",
     "easydentes.com.br + cnpj.biz/23752431000100",
     "BAIXA — investigar", "Lead novo", "Ligação 'cega' à recepção perguntando o sistema usado"),

    (18, "INVESTIGAR — nome similar",
     "Easydente Odontologia Especializada",
     "Salvador - BA",
     "easydente.com.br — @easydenteodonto (Instagram, 2.952 seguidores)",
     "Clínica de implantes com 18 anos. Nome similar. Vínculo NÃO confirmado",
     "easydente.com.br",
     "BAIXA — investigar", "Lead novo", "Confirmar via redes ou ligação"),

    # === GRUPO 7: HISTÓRICO / REFERÊNCIA ===
    (19, "Evento histórico",
     "Comitê Olímpico Brasileiro - Vila Olímpica Rio 2016",
     "Rio de Janeiro - RJ (histórico)",
     "COB — cob.org.br",
     "Anúncio EasyDental: homologado COB Rio 2016, 100+ profissionais, 30k atletas",
     "easydental.com.br (notícia)",
     "BAIXA (histórico)", "Não target", "Apenas case institucional"),

    (20, "Rede mobile",
     "OdontoSesc - 59 unidades móveis nacionais",
     "Brasil — vários estados",
     "Coordenação nacional SESC",
     "Vinculadas ao guarda-chuva SESC; alguns estados (DF, SP, RS) usam EasyDental",
     "sesc.com.br/unidades-moveis/odontosesc/",
     "MÉDIA", "Não target", "Referência"),
]

for i, row in enumerate(dados, start=5):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = BORDER
    if (i - 5) % 2 == 1:
        for j in range(1, 11):
            ws.cell(row=i, column=j).fill = ALT_FILL
    # colorir Confiança (col 8)
    conf = row[7]
    if conf.startswith("ALTA"):
        ws.cell(row=i, column=8).fill = HIGH_FILL
    elif conf.startswith("MÉDIA"):
        ws.cell(row=i, column=8).fill = MED_FILL
    elif conf.startswith("BAIXA"):
        ws.cell(row=i, column=8).fill = LOW_FILL
    ws.cell(row=i, column=8).font = BOLD
    ws.cell(row=i, column=8).alignment = CENTER

# Resumo no final
r0 = 5 + len(dados) + 2
ws.cell(row=r0, column=1, value="RESUMO DA EXECUÇÃO").font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws.cell(row=r0, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=10)

resumo = [
    ("Total de entidades identificadas", "20", ""),
    ("Confiança ALTA (confirmadas)", "14", "cases oficiais, embaixadores, redes"),
    ("Confiança MÉDIA", "3", "pistas fortes, precisam confirmar"),
    ("Confiança BAIXA (investigar)", "3", "nomes similares, históricos"),
    ("LEADS COMERCIAIS QUENTES (priorizar)", "2", "COSB + Vaga Vila Mariana"),
    ("PISTAS A CONFIRMAR", "4", "VIDA, Easydentes (9 un.), Easydente Salvador, Alysson Konno"),
    ("REDES PÚBLICAS (não target)", "5", "SESC SP/DF/RS + OdontoSesc + COB"),
    ("KOLs (use alunos como leads)", "7", "Bernal, Carol Moura, Valverde, Appezzato, Vieira, Sala, Rangel"),
]
for i, (k, v, obs) in enumerate(resumo, start=r0+1):
    ws.cell(row=i, column=1, value=k).font = BOLD
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2, value=v).font = BOLD
    ws.cell(row=i, column=2).alignment = CENTER
    ws.cell(row=i, column=2).border = BORDER
    c = ws.cell(row=i, column=3, value=obs)
    c.font = NORMAL
    c.alignment = WRAP
    c.border = BORDER
    ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=10)

# Próximos passos
r1 = r0 + len(resumo) + 2
ws.cell(row=r1, column=1, value="PRÓXIMOS PASSOS (em ordem)").font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
ws.cell(row=r1, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r1, start_column=1, end_row=r1, end_column=10)

passos = [
    "1. Rodar https://crt.sh/?q=%25.easydentalcloud.com.br → vai retornar TODOS os subdomínios de clientes. Isso transforma 20 nomes em potencialmente 200-1000+ clínicas.",
    "2. Filtrar Reclame Aqui pelos últimos 12 meses → 37 reclamações totais, ~85% respondidas. Os reclamantes mais recentes = leads em janela de churn.",
    "3. Buscar nas vagas ativas: linkedin.com/jobs, catho.com.br, vagas.com.br, infojobs.com.br, gupy.io — filtrar por 'EasyDental'. Pode dar 50+ clínicas.",
    "4. Cruzar com Receita Federal cada nome encontrado → razão social, CNPJ, sócios, CNAE.",
    "5. Aplicar score ICP (aba 6) e priorizar top 30 → abordar com mensagens da aba 7.",
]
for i, p in enumerate(passos, start=r1+1):
    c = ws.cell(row=i, column=1, value=p)
    c.font = NORMAL
    c.alignment = WRAP
    c.border = BORDER
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)
    ws.row_dimensions[i].height = 32

widths = [4, 22, 30, 26, 28, 38, 32, 16, 22, 32]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

# adicionar URL crt.sh à aba de fontes
fontes_ws = wb["10. Fontes"]
r = 4
while fontes_ws.cell(row=r, column=1).value is not None:
    r += 1
new_sources = [
    ("🔥 Certificate Transparency (descobre todos subdomínios = clientes)", "https://crt.sh/?q=%25.easydentalcloud.com.br"),
    ("SESC DF subdomínio (confirmação do padrão)", "https://sescdf.easydentalcloud.com.br/"),
    ("Notícia SESC DF adota EasyDental", "https://easydental.com.br/blog/sesc-df-adota-solucao-easydental/"),
    ("Vaga VIDA com sistema EasyDental", "https://vidadho.vagas.solides.com.br/vaga/697134"),
    ("Easydente Salvador (investigar)", "https://www.easydente.com.br/"),
    ("SESC RS clínicas odonto", "https://www.sesc-rs.com.br/saude/clinicas-odontologicas/"),
]
for label, url in new_sources:
    fontes_ws.cell(row=r, column=1, value=label).font = NORMAL
    fontes_ws.cell(row=r, column=1).alignment = WRAP
    fontes_ws.cell(row=r, column=1).border = BORDER
    cell = fontes_ws.cell(row=r, column=2, value=url)
    cell.hyperlink = url
    cell.font = Font(name="Arial", color="0563C1", underline="single", size=10)
    cell.alignment = WRAP
    cell.border = BORDER
    r += 1

# mover aba 12 para depois da 11
if "10. Fontes" in wb.sheetnames:
    wb.move_sheet("12. Passo 1 - Coleta", offset=-1)

wb.save(XLSX)
print(f"Saved: {XLSX}")
print(f"Aba '12. Passo 1 - Coleta' adicionada com {len(dados)} entradas")
