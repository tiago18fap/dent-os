from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ====== STYLES ======
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
SUB_FILL = PatternFill("solid", start_color="2E75B6")
ACCENT_FILL = PatternFill("solid", start_color="FFE699")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
WHITE_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="FFFFFF")
BOLD = Font(name="Arial", bold=True, size=10)
NORMAL = Font(name="Arial", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_title(ws, row, text, last_col, fill=HEADER_FILL):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = fill
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26

def apply_body(ws, start_row, end_row, last_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL
            cell.alignment = WRAP
            cell.border = BORDER
        if (r - start_row) % 2 == 1:
            for c in range(1, last_col + 1):
                ws.cell(row=r, column=c).fill = ALT_FILL

# ====== SHEET 1: Resumo Executivo ======
ws = wb.active
ws.title = "1. Resumo Executivo"

style_title(ws, 1, "Estratégia: Mapear Clínicas que usam EasyDental no Brasil", 3)

data = [
    ("Campo", "Conteúdo", "Observação"),
    ("Alvo", "EasyDental (Easy Software / OdontoPrev)", "Líder de mercado desde 1994, +45.000 dentistas no Brasil"),
    ("Versões ativas", "EasyDental Desktop (legado, em descontinuação) + EasyDental Cloud", "Desktop sendo migrado à força → fonte de insatisfação"),
    ("Domínios oficiais", "easydental.com.br, app.easydentalcloud.com.br, portal.easydentalcloud.com.br", "Use como base para Google dorks"),
    ("Casos-âncora públicos", "Rede COSB, SESC SP (35+ unidades), Comitê Olímpico (Rio 2016)", "Bons para case-study e benchmark, não para abordagem direta"),
    ("Dor identificada", "Migração forçada Desktop→Cloud com falhas, suporte ruim, perda de dados", "Reclame Aqui = principal mina de leads insatisfeitos"),
    ("Objetivo Vendii", "Identificar e abordar clínicas EasyDental para prospecção", ""),
    ("", "", ""),
    ("PIPELINE DE TRABALHO (em ordem)", "", ""),
    ("Etapa 1 — Coleta", "Rodar Google dorks (aba 2) + scraping Reclame Aqui (aba 3) + LinkedIn (aba 4)", "Output: lista bruta de razões sociais/clínicas"),
    ("Etapa 2 — Enriquecimento", "Cruzar com CRO, Receita Federal (CNPJ), Instagram, Google Maps", "Pegar telefone, e-mail, decisor (dentista responsável)"),
    ("Etapa 3 — Qualificação", "Filtrar por porte (≥2 cadeiras), região, tempo de uso, sinais de insatisfação", "Priorizar reclamantes recentes (<12 meses)"),
    ("Etapa 4 — Abordagem", "Outbound multicanal (WhatsApp + e-mail + LinkedIn) com pitch de migração", "Mensagem na aba 7"),
    ("Etapa 5 — Mensuração", "Tracking de tentativas, reuniões, conversão", "Template na aba 8"),
]

for i, row in enumerate(data, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 3)
apply_body(ws, 4, 3 + len(data) - 1, 3)

# Destacar título de pipeline
for r in [11]:
    for c in range(1, 4):
        ws.cell(row=r, column=c).fill = ACCENT_FILL
        ws.cell(row=r, column=c).font = BOLD

ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 48

# ====== SHEET 2: Google Dorks ======
ws = wb.create_sheet("2. Google Dorks")
style_title(ws, 1, "Google Dorks — comandos prontos para colar no Google", 3)

dorks = [
    ("#", "Dork (copie e cole no Google)", "O que retorna"),
    (1, '"easydental" site:.com.br -site:easydental.com.br', "Páginas .br que mencionam EasyDental fora do domínio oficial — geralmente clínicas falando do sistema"),
    (2, '"sistema easydental" clínica OR consultório', "Vagas, posts, depoimentos espontâneos de clínicas"),
    (3, '"conhecimento em easydental" OR "experiência com easydental"', "Vagas de recepcionista/secretária — clínica usa o sistema"),
    (4, 'site:linkedin.com/jobs "easydental"', "Vagas no LinkedIn que pedem EasyDental"),
    (5, 'site:catho.com.br "easydental"', "Vagas Catho que pedem EasyDental"),
    (6, 'site:vagas.com.br "easydental"', "Vagas vagas.com.br"),
    (7, 'site:indeed.com.br "easydental"', "Vagas Indeed BR"),
    (8, 'site:apcd.org.br "easydental"', "Classificados da APCD (Assoc. Paulista Cir.-Dentistas)"),
    (9, 'site:gupy.io "easydental"', "Vagas em ATS Gupy (redes maiores)"),
    (10, '"app.easydentalcloud.com.br" OR "portal.easydentalcloud.com.br"', "Páginas que linkam o app (clínicas usando)"),
    (11, '"easydental" "horário de atendimento" OR "agende online"', "Sites de clínicas integradas ao agendamento EasyDental"),
    (12, '"easydental" inurl:agendamento OR inurl:agenda', "Clínicas usando módulo de agendamento online"),
    (13, '"easydental cloud" filetype:pdf', "PDFs de cases, manuais, contratos vazados"),
    (14, '"easydental" "CNPJ"', "Razões sociais que aparecem junto à menção do sistema"),
    (15, 'site:reclameaqui.com.br "easy dental"', "Reclamações — leads insatisfeitos (alta intenção de migrar)"),
    (16, 'site:facebook.com "easydental" clínica', "Grupos/posts em FB"),
    (17, 'site:instagram.com "easydental"', "Stories e posts marcando o sistema"),
    (18, '"odontoprev" "easydental" parceria OR convênio', "Clínicas da rede OdontoPrev (dona do EasyDental) — alta probabilidade"),
    (19, '"easydental" rede OR franquia OR grupo odontológico', "Redes (alvo de maior ticket)"),
    (20, '"migrar do easydental" OR "sair do easydental" OR "trocar easydental"', "Intenção de churn explícita"),
    (21, '"easydental" depoimento OR avaliação OR review', "Reviews positivos e negativos"),
    (22, 'site:youtube.com "easydental" tutorial OR clínica', "Vídeos de clínicas mostrando o sistema (canal/descrição revela nome)"),
    (23, '"@easydental.com.br" OR "@easysoftware.com.br"', "E-mails ligados ao fornecedor"),
    (24, '"easydental" "São Paulo" OR "Rio de Janeiro" OR "Belo Horizonte" OR "Curitiba"', "Cortar por geografia — varie a cidade"),
    (25, 'site:cro-sp.org.br "easydental"', "Conselho Regional de Odontologia SP (repita para outros estados)"),
]

for i, row in enumerate(dorks, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 3)
apply_body(ws, 4, 3 + len(dorks) - 1, 3)

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 65
ws.column_dimensions['C'].width = 60

# ====== SHEET 3: Reclame Aqui (mina de ouro) ======
ws = wb.create_sheet("3. Reclame Aqui")
style_title(ws, 1, "Reclame Aqui — leads insatisfeitos (PRIORIDADE MÁXIMA)", 3)

ra_data = [
    ("Passo", "Ação", "Detalhe"),
    (1, "Acessar URL base", "https://www.reclameaqui.com.br/empresa/easy-dental/lista-reclamacoes/"),
    (2, "Filtrar por data", "Últimos 12 meses → quem está reclamando AGORA quer mudar"),
    (3, "Filtrar por status", '"Não resolvido" + "Não respondida" = clínica frustrada'),
    (4, "Extrair nome do reclamante", "Geralmente é a razão social, dentista responsável ou nome da clínica"),
    (5, "Buscar reclamante no Google", "Achar site, telefone, Instagram, CNPJ"),
    (6, "Validar porte", "Mínimo 2 cadeiras OU faturamento ≥ R$ 30k/mês (qualifica investimento)"),
    (7, "Abordagem empática", '"Vi sua reclamação sobre o EasyDental, posso te mostrar uma alternativa?"'),
    ("", "", ""),
    ("PADRÕES DE DOR JÁ MAPEADOS (use no pitch)", "", ""),
    ("Dor 1", "Migração Desktop→Cloud forçada e mal-feita (perda de dados, atrasos)", "Argumento: migração assistida sem perda"),
    ("Dor 2", "Cobrança de assinatura após já ter pago licença vitalícia", "Argumento: pricing transparente"),
    ("Dor 3", "Suporte demorado / sem resposta", "Argumento: SLA garantido"),
    ("Dor 4", "Cloud com menos relatórios financeiros que o Desktop", "Argumento: BI completo"),
    ("Dor 5", "Sistema não-intuitivo no Cloud", "Argumento: onboarding rápido + UX moderna"),
    ("Dor 6", "Recursos quebrados após migração (TISS, convênios, imagens)", "Argumento: integrações nativas funcionando"),
]

for i, row in enumerate(ra_data, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 3)
apply_body(ws, 4, 3 + len(ra_data) - 1, 3)

# Destacar título de seção
for c in range(1, 4):
    ws.cell(row=12, column=c).fill = ACCENT_FILL
    ws.cell(row=12, column=c).font = BOLD

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 60

# ====== SHEET 4: Canais OSINT ======
ws = wb.create_sheet("4. Canais OSINT")
style_title(ws, 1, "Canais de coleta — multi-fonte, em ordem de ROI", 4)

canais = [
    ("Canal", "URL/Como acessar", "Como usar", "ROI esperado"),
    ("Reclame Aqui", "reclameaqui.com.br/empresa/easy-dental/", "Filtrar últimos 12m, extrair nomes", "ALTO — intenção de churn explícita"),
    ("LinkedIn — vagas", "linkedin.com/jobs + filtro EasyDental", "Vagas mencionando EasyDental = clínica empregadora", "ALTO"),
    ("LinkedIn — perfis", 'busca: "EasyDental" + cargo recepcionista/secretária odontológica', "Perfil revela onde a pessoa trabalha", "ALTO"),
    ("Catho / Vagas.com / Indeed / Gupy", "Buscar EasyDental nas plataformas", "Mesma lógica das vagas LinkedIn, mais clínicas pequenas", "MÉDIO-ALTO"),
    ("APCD Classificados", "apcd.org.br/classificados-da-apcd/empregos", "Vagas em SP que pedem EasyDental", "MÉDIO"),
    ("CRO Classificados", "cro-rj.org.br/classificados (repetir p/ outros estados)", "Listagens regionais", "MÉDIO"),
    ("Google dorks (aba 2)", "google.com", "Rodar os 25 dorks da aba 2", "ALTO"),
    ("Instagram (hashtag/local)", "instagram.com + #easydental #gestaoodontologica", "Posts de clínicas mostrando print do sistema", "MÉDIO"),
    ("Facebook Grupos", "Grupos: 'Gestão de Clínica Odonto', 'Dentistas Brasil'", "Buscas internas por EasyDental", "MÉDIO"),
    ("YouTube", "Vídeos de tutorial EasyDental — descrição revela clínica", "Identificar canais que mostram a tela", "BAIXO-MÉDIO"),
    ("Reddit / Comunidades", "r/Odontologia, fóruns OdontoBrasil", "Discussões sobre sistemas", "BAIXO"),
    ("Eventos (CIOSP, CIORJ)", "Lista de expositores e patrocinadores", "EasyDental expõe → quem visita o estande é cliente/prospect", "MÉDIO (uma vez/ano)"),
    ("Parceiros EasyDental", "easydental.com.br/parceiros/", "Lista de revendas/consultorias parceiras = atendem clientes EasyDental", "MÉDIO (indireto)"),
    ("OdontoPrev (dona)", "odontoprev.com.br rede credenciada", "Cruzar credenciados OdontoPrev com EasyDental (alta correlação)", "MÉDIO"),
    ("Wayback Machine", "web.archive.org de sites de clínicas", "Histórico revela mudanças/integrações com EasyDental", "BAIXO"),
    ("BuiltWith / Wappalyzer", "builtwith.com / wappalyzer.com", "Detectar pixels, scripts (raro p/ ERP, mas vale tentar agenda online)", "BAIXO"),
]

for i, row in enumerate(canais, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 4)
apply_body(ws, 4, 3 + len(canais) - 1, 4)

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 22

# ====== SHEET 5: Enriquecimento ======
ws = wb.create_sheet("5. Enriquecimento")
style_title(ws, 1, "Enriquecimento — de nome bruto a lead qualificado", 3)

enr = [
    ("Dado a buscar", "Onde encontrar", "Ferramenta sugerida"),
    ("CNPJ + Razão Social", "Receita Federal / cnpj.ws / casadosdados.com.br", "API casadosdados ou consulta manual"),
    ("Endereço completo", "Google Maps + Receita Federal", "Maps API"),
    ("Telefone fixo / WhatsApp", "Site da clínica, Google Maps, Instagram bio", "Manual + scraping"),
    ("E-mail", "Site da clínica (rodapé/contato), Hunter.io, Apollo", "Hunter.io / Apollo / Snov.io"),
    ("Dentista responsável (decisor)", "CRO do estado + LinkedIn + site", "Busca CRO público"),
    ("CNAE / porte / nº funcionários", "Receita Federal + LinkedIn 'Pessoas'", "casadosdados.com.br"),
    ("Convênios aceitos", "Site da clínica", "Indica perfil (OdontoPrev = alta correlação EasyDental)"),
    ("Tempo de uso do EasyDental", "Depoimentos, posts antigos, Wayback", "Indica maturidade do lock-in"),
    ("Volume de atendimento", "Google Maps reviews + Instagram seguidores", "Proxy de porte"),
    ("Sócios e histórico", "Receita Federal — Quadro Societário", "casadosdados.com.br"),
    ("Score de prioridade", "Calcular: (Dor x Porte x Recência reclamação)", "Coluna calculada na aba 9"),
]

for i, row in enumerate(enr, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 3)
apply_body(ws, 4, 3 + len(enr) - 1, 3)

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 38

# ====== SHEET 6: Qualificação (ICP) ======
ws = wb.create_sheet("6. ICP e Qualificação")
style_title(ws, 1, "ICP — perfil ideal de clínica EasyDental para abordagem", 3)

icp = [
    ("Critério", "Faixa ideal", "Por quê"),
    ("Porte (cadeiras odontológicas)", "≥ 2", "Investimento em ERP só faz sentido a partir disso"),
    ("Faturamento estimado", "≥ R$ 30k/mês", "Capacidade de pagar mensalidade premium"),
    ("Tempo no EasyDental", "≥ 2 anos OU recém-migrado ao Cloud", "Lock-in maduro OU acabou de sofrer migração ruim"),
    ("Localização", "Capitais e regiões metropolitanas (SP, RJ, MG, PR, RS, DF) — fase 1", "Densidade + ticket maior"),
    ("Especialidade", "Clínica geral, ortodontia, implante, estética", "Maior fluxo de pacientes = maior dor em gestão"),
    ("Sinais de insatisfação", "Reclame Aqui, posts negativos, churn de funcionários", "Janela quente de abordagem"),
    ("Decisor identificado", "Dentista-proprietário OU gestor administrativo", "Sem decisor mapeado, não avançar"),
    ("Está na rede OdontoPrev", "Sim/Não", "Sim aumenta probabilidade de já ter EasyDental"),
    ("Já automatiza WhatsApp/agenda", "Sim", "Indica abertura a tecnologia"),
    ("", "", ""),
    ("SCORE DE PRIORIDADE (0-100)", "", ""),
    ("Reclamação Reclame Aqui últimos 6 meses", "+30 pts", ""),
    ("Reclamação Reclame Aqui 7-12 meses", "+15 pts", ""),
    ("≥ 3 cadeiras / rede", "+20 pts", ""),
    ("Capital ou região metropolitana top 6", "+10 pts", ""),
    ("Decisor identificado com contato", "+15 pts", ""),
    ("Insatisfação pública (post, review)", "+15 pts", ""),
    ("Cliente OdontoPrev", "+10 pts", ""),
]

for i, row in enumerate(icp, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 3)
apply_body(ws, 4, 3 + len(icp) - 1, 3)

# Destaca subseção
for c in range(1, 4):
    ws.cell(row=14, column=c).fill = ACCENT_FILL
    ws.cell(row=14, column=c).font = BOLD

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 50

# ====== SHEET 7: Mensagens de Abordagem ======
ws = wb.create_sheet("7. Mensagens")
style_title(ws, 1, "Templates de abordagem multicanal", 2)

msgs = [
    ("Canal", "Mensagem"),
    ("WhatsApp — pós Reclame Aqui",
     "Oi Dr(a). [NOME], tudo bem? Vi seu relato no Reclame Aqui sobre o EasyDental "
     "(migração / suporte / etc). Trabalho com gestão odontológica e ajudo clínicas como a sua "
     "a migrar sem perder dado e sem ficar parado. Posso te mostrar uma demo de 15 min? — [SEU NOME], Vendii"),
    ("WhatsApp — frio",
     "Oi Dr(a). [NOME], vi que a [CLÍNICA] usa EasyDental. Estamos lançando uma ferramenta de "
     "[posicionamento Vendii em 1 linha] que integra com o que você já tem e tira fricção do "
     "dia-a-dia. Topa 15 min nesta semana? — [SEU NOME]"),
    ("E-mail — assunto",
     "Migração sem dor do EasyDental — [CLÍNICA]"),
    ("E-mail — corpo",
     "Dr(a). [NOME],\n\n"
     "Reparei que a [CLÍNICA] usa EasyDental há [TEMPO]. Tenho falado com várias clínicas "
     "passando pelo mesmo desafio: o Desktop está sendo descontinuado, o Cloud trouxe mudanças "
     "no fluxo financeiro e o suporte vem deixando a desejar.\n\n"
     "Em [TEMPO] de implementação, conseguimos:\n"
     "• Manter 100% do histórico de pacientes\n"
     "• Integrar agenda + WhatsApp sem retrabalho\n"
     "• [PROVA SOCIAL: case parecido]\n\n"
     "Posso te mostrar em 15 min? Sugestão de horário: [DATA].\n\n"
     "Abraço,\n[NOME] — Vendii"),
    ("LinkedIn — conexão",
     "Dr(a). [NOME], converso com clínicas que usam EasyDental e enfrentam o cenário de "
     "descontinuação do Desktop. Queria trocar uma ideia rápida com você."),
    ("LinkedIn — InMail",
     "Dr(a). [NOME], a [CLÍNICA] aparece no nosso radar de clínicas EasyDental no Brasil. "
     "Estamos ajudando gestores a sair da migração forçada sem perder produtividade. "
     "Posso te mandar um material de 2 páginas comparando alternativas? — Vendii"),
    ("Follow-up #1 (D+3)",
     "Dr(a). [NOME], só voltando aqui caso minha primeira mensagem tenha passado batido. "
     "Sem compromisso — separo 15 min na sua semana?"),
    ("Follow-up #2 (D+7)",
     "Dr(a). [NOME], última tentativa por aqui. Se preferir, mando um vídeo curto de 2 min "
     "mostrando como funciona — me responde só 'manda' que eu envio."),
    ("Quebra de objeção — 'já estou no Cloud'",
     "Entendo, Dr(a). Vários clientes nossos chegaram do EasyDental Cloud justamente por "
     "[dor X]. Se quiser, te mando um comparativo lado-a-lado e você decide se vale 15 min."),
    ("Quebra de objeção — 'sem tempo'",
     "Sem problemas. Posso mandar 1 print + 1 vídeo de 90s; em 2 min você entende e decide."),
]

for i, row in enumerate(msgs, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 2)
apply_body(ws, 4, 3 + len(msgs) - 1, 2)

ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 90

# ajustar altura
for r in range(4, 3 + len(msgs)):
    ws.row_dimensions[r].height = 90

# ====== SHEET 8: Tracking de Leads ======
ws = wb.create_sheet("8. Tracking de Leads")
style_title(ws, 1, "Pipeline de leads — preencher conforme prospecção avança", 14)

headers = [
    "ID", "Clínica / Razão Social", "CNPJ", "Cidade", "UF", "Decisor",
    "Telefone/WhatsApp", "E-mail", "Fonte (Reclame Aqui/LinkedIn/etc)",
    "Score Prioridade", "Status", "Última ação", "Próximo passo",
    "Data próxima ação"
]
for j, h in enumerate(headers, start=1):
    ws.cell(row=3, column=j, value=h)

# Linha de exemplo
example = [
    1, "Clínica Exemplo Ltda", "00.000.000/0001-00", "São Paulo", "SP",
    "Dr. João Silva", "(11) 9XXXX-XXXX", "contato@exemplo.com.br",
    "Reclame Aqui — migração", 75, "1 - Pesquisado", "Envio WhatsApp",
    "Follow-up D+3", "2026-06-05"
]
for j, v in enumerate(example, start=1):
    ws.cell(row=4, column=j, value=v)

# 50 linhas em branco com bordas para preencher
for r in range(5, 55):
    for c in range(1, 15):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).font = NORMAL

style_header(ws, 3, 14)

# tornar exemplo destacado
for c in range(1, 15):
    ws.cell(row=4, column=c).fill = ACCENT_FILL
    ws.cell(row=4, column=c).font = NORMAL
    ws.cell(row=4, column=c).border = BORDER

widths = [5, 30, 22, 16, 5, 22, 22, 26, 28, 12, 18, 28, 24, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "B4"

# Resumo no topo do sheet (acima do header)
ws.cell(row=2, column=1, value="Total leads:").font = BOLD
ws.cell(row=2, column=2, value='=COUNTA(B4:B1000)-1')

# ====== SHEET 9: KPIs ======
ws = wb.create_sheet("9. KPIs")
style_title(ws, 1, "KPIs de prospecção — metas semanais", 3)

kpi = [
    ("KPI", "Meta semanal", "Cálculo"),
    ("Novos leads identificados (brutos)", 50, ""),
    ("Leads enriquecidos (com decisor + contato)", 25, ""),
    ("Leads qualificados (score ≥ 60)", 12, ""),
    ("Tentativas de contato", 60, "≈ 5 toques por lead qualificado"),
    ("Respostas", 6, "≈ 10% taxa de resposta"),
    ("Reuniões agendadas", 3, "≈ 50% das respostas viram reunião"),
    ("Reuniões realizadas", 2, "≈ 65% no-show"),
    ("Propostas enviadas", 1, ""),
    ("Fechamentos/mês (extrapolado)", "1-2", "Funil semanal x 4"),
    ("", "", ""),
    ("FUNIL ESPERADO (12 semanas)", "", ""),
    ("Leads brutos coletados", 600, ""),
    ("Qualificados", 144, ""),
    ("Reuniões realizadas", 24, ""),
    ("Clientes fechados", "6-10", ""),
]

for i, row in enumerate(kpi, start=3):
    for j, val in enumerate(row, start=1):
        ws.cell(row=i, column=j, value=val)

style_header(ws, 3, 3)
apply_body(ws, 4, 3 + len(kpi) - 1, 3)

# Destaca subseção
for c in range(1, 4):
    ws.cell(row=14, column=c).fill = ACCENT_FILL
    ws.cell(row=14, column=c).font = BOLD

ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 45

# ====== SHEET 10: Fontes e Links ======
ws = wb.create_sheet("10. Fontes")
style_title(ws, 1, "Fontes oficiais e referências (clicáveis)", 2)

fontes = [
    ("Categoria", "Link"),
    ("Site oficial EasyDental", "https://easydental.com.br/"),
    ("Quem somos / Histórico", "https://easydental.com.br/quem-somos/"),
    ("Depoimentos públicos", "https://easydental.com.br/depoimentos/"),
    ("Estudos de caso (Blog)", "https://easydental.com.br/blog/estudos-de-casos/veja-os-depoimentos-dos-nossos-clientes-sobre-o-easydental-cloud/"),
    ("Lista de parceiros", "https://easydental.com.br/parceiros/"),
    ("App login (subdomínio)", "https://app.easydentalcloud.com.br/"),
    ("Portal Cloud", "https://portal.easydentalcloud.com.br/"),
    ("LinkedIn EasyDental", "https://br.linkedin.com/company/easydental"),
    ("LinkedIn EasyDental Cloud", "https://br.linkedin.com/company/easydental-cloud"),
    ("Reclame Aqui — lista reclamações", "https://www.reclameaqui.com.br/empresa/easy-dental/lista-reclamacoes/"),
    ("Reclame Aqui — migração forçada (case)", "https://www.reclameaqui.com.br/easy-dental/cobranca-de-assinatura-e-migracao-forcada-para-nuvem-do-software-easy-dental_9ti5TRDwlrkfSvXD/"),
    ("APCD Classificados (vagas SP)", "https://www.apcd.org.br/classificados-da-apcd/empregos"),
    ("CRO-RJ Classificados", "https://classificados.cro-rj.org.br/list"),
    ("OdontoPrev (dona do EasyDental desde 2008)", "https://www.odontoprev.com.br/"),
    ("Easy Dental US (Henry Schein — descontinuado 2023)", "https://www.easydental.com/"),
]

for i, row in enumerate(fontes, start=3):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        if j == 2 and i > 3 and val.startswith("http"):
            cell.hyperlink = val
            cell.font = Font(name="Arial", color="0563C1", underline="single", size=10)

style_header(ws, 3, 2)
apply_body(ws, 4, 3 + len(fontes) - 1, 2)
# refazer links (apply_body sobrescreveu font)
for i in range(4, 3 + len(fontes)):
    val = ws.cell(row=i, column=2).value
    if isinstance(val, str) and val.startswith("http"):
        ws.cell(row=i, column=2).hyperlink = val
        ws.cell(row=i, column=2).font = Font(name="Arial", color="0563C1", underline="single", size=10)

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 90

# Save
out = "/sessions/dreamy-brave-keller/mnt/GERADO PELO COWORK/EasyDental.xlsx"
wb.save(out)
print(f"Saved: {out}")