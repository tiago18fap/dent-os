"""Adiciona aba '11. Clínicas Confirmadas' com nomes reais coletados via OSINT."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/sessions/dreamy-brave-keller/mnt/GERADO PELO COWORK/EasyDental.xlsx"
wb = load_workbook(XLSX)

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
HIGH_FILL = PatternFill("solid", start_color="C6EFCE")
MED_FILL = PatternFill("solid", start_color="FFEB9C")
LOW_FILL = PatternFill("solid", start_color="FFC7CE")
WHITE_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="FFFFFF")
BOLD = Font(name="Arial", bold=True, size=10)
NORMAL = Font(name="Arial", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINK_FONT = Font(name="Arial", color="0563C1", underline="single", size=10)

if "11. Clínicas Confirmadas" in wb.sheetnames:
    del wb["11. Clínicas Confirmadas"]
ws = wb.create_sheet("11. Clínicas Confirmadas")

# Título
ws.merge_cells("A1:I1")
c = ws["A1"]
c.value = "Clínicas e dentistas que usam EasyDental — coletado via OSINT (público)"
c.fill = HEADER_FILL
c.font = TITLE_FONT
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

# Nota
ws.merge_cells("A2:I2")
c = ws["A2"]
c.value = ("Fontes: site oficial easydental.com.br (depoimentos, parceiros, blog/cases), Reclame Aqui, "
           "LinkedIn, Receita Federal, Google Maps. CONFIRMAR vínculo ativo antes de abordar.")
c.font = Font(name="Arial", italic=True, size=9, color="595959")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws.row_dimensions[2].height = 28

headers = ["#", "Tipo", "Nome / Razão Social", "Decisor / Responsável", "Cidade-UF",
           "Contato", "Como foi identificado (evidência pública)", "Confiança", "Próxima ação sugerida"]
for j, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=j, value=h)
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = BORDER

# Dados confirmados — nomes coletados via OSINT público
clinicas = [
    # CASES OFICIAIS (alta confiança)
    (1, "Rede / Cliente case oficial",
     "Myai Odontologia (Odonto Myai)",
     "Dr. Edson Myai (CRO-SP 59.180)",
     "São Paulo - SP (Vila Antonieta / Aricanduva)",
     "(11) 2727-4743 — odontomyai.com.br — @edson.myai",
     "Estudo de caso publicado no blog oficial EasyDental: 'Implantação do EasyDental na Clínica Myai'. Cliente desde 2012. 15+ dentistas, 3 consultórios.",
     "ALTA", "Abordagem de upsell/migração — cliente maduro, candidato a churn por insatisfação ou expansão"),

    (2, "Rede multi-unidades",
     "COSB - Centro de Ortodontia e Saúde Bucal",
     "Diretoria COSB",
     "São Bernardo do Campo e Diadema - SP",
     "(11) 4122-0550 — cosb.com.br — CNPJ 01.740.990/0001-67",
     "Depoimento oficial no site EasyDental: 'usa EasyDental desde 2012, em fase avançada de migração para o Cloud'. 6 unidades no ABC. Planeja operação nos EUA.",
     "ALTA", "Conta-chave (rede). Mapear comitê de TI. Pitch: BI consolidado multi-unidade"),

    (3, "Rede pública / governo",
     "SESC São Paulo",
     "Gerência de Odontologia SESC-SP",
     "Estado de SP - 34+ unidades",
     "sescsp.org.br — licitações: licitacoes-e-contratacoes",
     "Anúncio oficial EasyDental: 'após licitação, SESC SP adotou EasyDental para gestão odontológica' (2017). 34+ unidades.",
     "ALTA", "Rede pública — não é prospect comercial direto (compra via licitação)"),

    (4, "Evento / institucional histórico",
     "Comitê Olímpico Brasileiro (COB) — Vila Olímpica Rio 2016",
     "COB",
     "Rio de Janeiro - RJ (histórico)",
     "cob.org.br",
     "Anúncio EasyDental: 'homologado pelo COB para gerir clínicas odontológicas Rio 2016, 100+ profissionais, 30 mil atletas'.",
     "BAIXA (histórico)", "Apenas referência/case — não target"),

    # DEPOIMENTOS PÚBLICOS NO SITE
    (5, "Consultório individual — depoimento oficial",
     "Consultório Dr. Nelson Dall'Oca",
     "Dr. Nelson Alves Dall'Oca (cirurgião-dentista)",
     "São Paulo - SP",
     "ident.com.br/dalloca — doctoralia.com.br/nelson-dall-oca",
     "Depoimento oficial site EasyDental: 'uso há anos, me ajuda a focar na parte técnica'.",
     "ALTA", "Lead morno — relação longa com EasyDental, mas elegível a comparação se houver dor"),

    # PARCEIROS / EMBAIXADORES (todos clientes ativos divulgando o sistema)
    (6, "Embaixador / Parceiro",
     "Instituto Bernal de Odontologia",
     "Dr. Anderson (Aderson) Bernal",
     "São Paulo - SP",
     "institutobernal.com.br — andersonbernal.com.br — @institutobernal",
     "Página oficial de parceiros EasyDental. Diretor/fundador do Instituto Bernal, referência em odontologia digital (CEREC CAD/CAM).",
     "ALTA", "Parceiro estratégico do concorrente — abordagem indireta (via aluno/seguidor)"),

    (7, "Embaixador / Parceiro",
     "Clínica / Curso Dra. Carol Moura",
     "Dra. Carol Moura",
     "Brasil",
     "Buscar no Instagram @carolmoura (dentista)",
     "Página oficial de parceiros EasyDental: 'oferece desconto exclusivo no EasyDental para alunos'.",
     "ALTA", "Embaixadora do concorrente — não prospect direto; mapear alunos dela como leads"),

    (8, "Embaixador / Parceiro",
     "Curso / Clínica Dr. Felipe Valverde",
     "Dr. Felipe Valverde",
     "Brasil",
     "Buscar no Instagram",
     "Página parceiros EasyDental — desconto para alunos.",
     "ALTA", "Mesma lógica — alunos dele são target"),

    (9, "Embaixador / Parceiro",
     "Clínica Dr. André Appezzato",
     "Dr. André Appezzato",
     "Brasil",
     "Buscar Instagram/LinkedIn",
     "Página parceiros EasyDental.",
     "ALTA", "Parceiro pago do EasyDental — não target direto"),

    (10, "Embaixador / Parceiro",
     "Clínica Dr. Rodrigo Vieira",
     "Dr. Rodrigo Vieira",
     "Brasil",
     "Buscar Instagram/LinkedIn",
     "Página parceiros EasyDental.",
     "ALTA", "Parceiro pago — não target direto"),

    (11, "Embaixador / Parceiro",
     "Clínica Dr. Felipe Sala",
     "Dr. Felipe Sala",
     "Brasil",
     "Buscar Instagram/LinkedIn",
     "Página parceiros EasyDental.",
     "ALTA", "Parceiro pago — não target direto"),

    (12, "Embaixador / Parceiro",
     "Clínica Dr. Rafael Rangel",
     "Dr. Rafael Rangel ('Rafa Rangel')",
     "Brasil",
     "@dr.rafaelrangel (Facebook/Instagram)",
     "Página parceiros EasyDental.",
     "ALTA", "Parceiro pago — não target direto"),

    (13, "Cliente depoente",
     "Consultório Dr. Alysson Konno",
     "Dr. Alysson Konno (cirurgião-dentista)",
     "(Brasil — confirmar UF)",
     "Buscar CRO + LinkedIn",
     "Depoimento no site EasyDental sobre o suporte e versão Cloud.",
     "MÉDIA", "Cliente satisfeito — abordagem comparativa apenas se houver dor"),

    # POSSÍVEL CLIENTE (nome similar — investigar)
    (14, "INVESTIGAR — nome similar",
     "Easydentes Clínicas Odontológicas",
     "CNPJ 23.752.431/0001-00",
     "São Paulo - SP (9 unidades: Itaim Paulista, Sapopemba, Sacomã, V. Mercês, Av. Oratório, etc.)",
     "easydentes.com.br",
     "Rede com 9 unidades e nome similar ao sistema. ALTA correlação mas vínculo NÃO confirmado em fontes oficiais. Pode ser parceira ou só coincidência de nome.",
     "BAIXA — investigar", "Confirmar via LinkedIn de funcionários ou ligação 'cega' à recepção perguntando o sistema usado"),

    # PISTAS DERIVADAS — clínicas a investigar a partir de vagas/reclamações
    (15, "Pista — vaga em SP (Vila Mariana)",
     "Clínica não identificada (Vila Mariana, SP)",
     "—",
     "São Paulo - SP (Vila Mariana)",
     "Catho — buscar vaga arquivada",
     "Vaga de recepcionista CLT pedindo 'conhecimento no sistema EasyDental será um diferencial' (referência em busca pública).",
     "MÉDIA", "Buscar vaga ativa/arquivada na Catho/InfoJobs/Vagas.com e identificar contratante"),

    # REDE OdontoPrev (dona) — leads indiretos
    (16, "Pista — alta correlação",
     "Clínicas credenciadas OdontoPrev",
     "Vários",
     "Brasil (rede nacional)",
     "odontoprev.com.br/rede",
     "OdontoPrev é dona do Easy Software desde 2008. Há forte sobreposição entre rede credenciada e usuários EasyDental.",
     "MÉDIA-ALTA", "Baixar lista de credenciados por região e cruzar com perfil ICP"),
]

for i, row in enumerate(clinicas, start=4):
    for j, val in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = NORMAL
        cell.alignment = WRAP
        cell.border = BORDER
    # zebra
    if (i - 4) % 2 == 1:
        for j in range(1, 10):
            ws.cell(row=i, column=j).fill = ALT_FILL
    # colorir coluna Confiança (H = 8)
    conf = row[7]
    fill = None
    if conf.startswith("ALTA"):
        fill = HIGH_FILL
    elif conf.startswith("MÉDIA"):
        fill = MED_FILL
    elif conf.startswith("BAIXA"):
        fill = LOW_FILL
    if fill:
        ws.cell(row=i, column=8).fill = fill
        ws.cell(row=i, column=8).font = BOLD
        ws.cell(row=i, column=8).alignment = CENTER

widths = [5, 28, 32, 28, 28, 32, 50, 18, 38]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A4"

# legenda
last_row = 4 + len(clinicas) + 2
ws.cell(row=last_row, column=1, value="Legenda confiança:").font = BOLD
ws.cell(row=last_row, column=2, value="ALTA = fonte oficial confirma uso atual ou recente")
ws.cell(row=last_row+1, column=2, value="MÉDIA = sinal forte mas precisa confirmar vínculo ativo")
ws.cell(row=last_row+2, column=2, value="BAIXA = pista indireta, investigar antes de abordar")
for r in range(last_row, last_row+3):
    ws.cell(row=r, column=2).font = NORMAL
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center")

# mover aba para penúltima posição (antes de Fontes)
order = wb.sheetnames
if "10. Fontes" in order:
    wb.move_sheet("11. Clínicas Confirmadas", offset=-1)

# adicionar fontes novas ao sheet 10
fontes_ws = wb["10. Fontes"]
add = [
    ("Case oficial Myai (blog EasyDental)", "https://easydental.com.br/blog/automacao-de-processos/estudo-de-caso-implantacao-do-easydental-na-clinica-myai/"),
    ("Site Odonto Myai (cliente)", "https://odontomyai.com.br/"),
    ("Site COSB (cliente)", "https://www.cosb.com.br/"),
    ("Notícia oficial SESC-SP adota EasyDental", "https://www.easydental.com.br/noticias/99-sesc-sao-paulo-adota-solucao-da-easydental-do-grupo-odontoprev-para-a-gestao-de-seu-servico-de-odontologia"),
    ("Site Dr. Anderson Bernal (parceiro)", "https://www.andersonbernal.com.br/"),
    ("Instituto Bernal (parceiro)", "https://www.instagram.com/institutobernal/"),
    ("Dr. Nelson Dall'Oca - iDent", "https://www.ident.com.br/dalloca"),
    ("Easydentes - CNPJ (investigar)", "https://cnpj.biz/23752431000100"),
]
# achar primeira linha vazia
r = 4
while fontes_ws.cell(row=r, column=1).value is not None:
    r += 1
for label, url in add:
    fontes_ws.cell(row=r, column=1, value=label).font = NORMAL
    fontes_ws.cell(row=r, column=1).alignment = WRAP
    fontes_ws.cell(row=r, column=1).border = BORDER
    cell = fontes_ws.cell(row=r, column=2, value=url)
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.alignment = WRAP
    cell.border = BORDER
    r += 1

wb.save(XLSX)
print(f"Saved: {XLSX}")
print(f"Aba '11. Clínicas Confirmadas' adicionada com {len(clinicas)} entradas")
